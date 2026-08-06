from __future__ import annotations

import json
import tempfile
import unittest
from datetime import date, datetime, timezone
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import Workbook

from app.clients.http_ozon_seller import (
    OzonSellerClient,
    create_placement_by_products_report,
    create_placement_by_supplies_report,
    fetch_analytics_stocks,
    fetch_product_info_list,
    fetch_report_info,
    iter_product_list,
)
from app.jobs.job_ozon_placement import default_placement_report_date
from app.clients.local_source_files import (
    read_first_sheet_rows,
    read_local_table_rows,
    read_production_inventory_rows,
    read_supply_pipeline_rows,
    read_tabular_text,
    resolve_latest_file,
)
from app.normalize.norm_ozon_placement import parse_placement_xlsx
from app.ops.ozon_placement_reports import download_placement_by_supplies_report
from app.normalize.norm_ozon_stocks import (
    merge_analytics_stock_rows,
    normalize_analytics_stock,
    normalize_product_info_item,
    normalize_product_list_item,
)


class OzonStocksClientTests(unittest.TestCase):
    def test_product_list_and_stock_methods(self) -> None:
        session = Mock()
        payloads = [
            {"result": {"items": [{"product_id": 1}], "last_id": ""}},
            {"items": [{"id": 1, "offer_id": "ART", "sku": 10}]},
            {"items": [{"sku": 10, "offer_id": "ART", "available_stock_count": 5}]},
        ]
        responses = []
        for payload in payloads:
            response = Mock(status_code=200, content=json.dumps(payload).encode("utf-8"))
            response.json.return_value = payload
            responses.append(response)
        session.post.side_effect = responses
        client = OzonSellerClient(client_id="client", api_key="key", session=session)

        products = list(iter_product_list(client))
        info, _ = fetch_product_info_list(client, product_ids=[1])
        stocks, _ = fetch_analytics_stocks(client, skus=[10])

        self.assertEqual(products[0][0][0]["product_id"], 1)
        self.assertEqual(info[0]["offer_id"], "ART")
        self.assertEqual(stocks[0]["available_stock_count"], 5)

    def test_report_methods(self) -> None:
        session = Mock()
        payloads = [{"code": "abc"}, {"result": {"status": "success", "file": "https://example.test/report.xlsx"}}]
        responses = []
        for payload in payloads:
            response = Mock(status_code=200, content=json.dumps(payload).encode("utf-8"))
            response.json.return_value = payload
            responses.append(response)
        session.post.side_effect = responses
        client = OzonSellerClient(client_id="client", api_key="key", session=session)

        code, _ = create_placement_by_products_report(client, date_from=date(2026, 7, 1), date_to=date(2026, 7, 1))
        info, _ = fetch_report_info(client, code=code)

        self.assertEqual(code, "abc")
        self.assertEqual(info["status"], "success")

    def test_placement_by_supplies_report_method(self) -> None:
        session = Mock()
        payload = {"code": "supplies-code"}
        response = Mock(status_code=200, content=json.dumps(payload).encode("utf-8"))
        response.json.return_value = payload
        session.post.return_value = response
        client = OzonSellerClient(client_id="client", api_key="key", session=session)

        code, log = create_placement_by_supplies_report(client, date_from=date(2026, 8, 5), date_to=date(2026, 8, 5))

        self.assertEqual(code, "supplies-code")
        self.assertEqual(log.method_name, "ozon_placement_by_supplies_create")
        self.assertEqual(session.post.call_args.args[0], "https://api-seller.ozon.ru/v1/report/placement/by-supplies/create")
        self.assertEqual(
            session.post.call_args.kwargs["json"],
            {"date_from": "2026-08-05", "date_to": "2026-08-05"},
        )


class OzonStocksNormalizationTests(unittest.TestCase):
    def test_normalize_product_and_stock_rows(self) -> None:
        product = normalize_product_list_item({"product_id": "1", "offer_id": "ART", "sku": "10", "archived": False})
        info = normalize_product_info_item(
            {
                "id": "1",
                "offer_id": "ART",
                "sku": "10",
                "name": "Name",
                "price": "123,45",
                "currency_code": "RUB",
                "statuses": {"status": "price_sent", "status_name": "Готов к продаже"},
                "visibility_details": {"has_price": True, "has_stock": False},
                "model_info": {"model_id": "99", "count": "2"},
                "images": ["https://example.test/1.jpg"],
                "barcodes": ["123"],
            }
        )
        stock = normalize_analytics_stock(
            {
                "sku": "10",
                "offer_id": "ART",
                "cluster_id": "2",
                "warehouse_id": "100",
                "warehouse_name": "WAREHOUSE",
                "placement_zone": "SORT",
                "available_stock_count": "5",
                "valid_stock_count": "4",
                "other_stock_count": "1",
                "requested_stock_count": "1",
                "transit_stock_count": "2",
                "return_from_customer_stock_count": "3",
                "return_to_seller_stock_count": "4",
                "stock_defect_stock_count": "5",
                "transit_defect_stock_count": "6",
                "expiring_stock_count": "7",
                "waiting_docs_stock_count": "8",
                "waiting_docs_to_export_stock_count": "9",
                "excess_stock_count": "10",
            }
        )
        self.assertEqual(product["product_id"], 1)
        self.assertEqual(info["sku"], 10)
        self.assertEqual(info["price"], 123.45)
        self.assertEqual(info["status_name"], "Готов к продаже")
        self.assertEqual(info["has_price"], True)
        self.assertEqual(info["model_id"], 99)
        self.assertEqual(info["images_count"], 1)
        self.assertEqual(stock["in_way_to_warehouse_count"], 6)
        self.assertEqual(stock["warehouse_name"], "WAREHOUSE")
        self.assertEqual(stock["valid_stock_count"], 4)
        self.assertEqual(stock["waiting_docs_to_export_stock_count"], 9)

    def test_merge_duplicate_stock_rows_for_staging(self) -> None:
        rows = [
            {"sku": 10, "offer_id": "ART", "cluster_id": 1, "available_stock_count": 2, "requested_stock_count": 1, "transit_stock_count": 0, "return_from_customer_stock_count": 0, "in_way_to_warehouse_count": 1, "payload": {"row": 1}},
            {"sku": 10, "offer_id": "ART", "cluster_id": 1, "available_stock_count": 3, "requested_stock_count": 0, "transit_stock_count": 2, "return_from_customer_stock_count": 0, "in_way_to_warehouse_count": 2, "payload": {"row": 2}},
        ]
        merged = merge_analytics_stock_rows(rows)
        self.assertEqual(len(merged), 1)
        self.assertEqual(merged[0]["available_stock_count"], 5)
        self.assertEqual(merged[0]["in_way_to_warehouse_count"], 3)


class PlacementAndLocalFileTests(unittest.TestCase):
    def test_default_placement_report_date_uses_minsk_current_day(self) -> None:
        value = default_placement_report_date(datetime(2026, 8, 6, 4, 0, tzinfo=timezone.utc))

        self.assertEqual(value, date(2026, 8, 6))

    def test_parse_placement_xlsx(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["SKU", "Артикул", "Название", "Стоимость размещения"])
        sheet.append([10, "ART", "Name", "12,5"])
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            xlsx_path = Path(tmp) / "placement.xlsx"
            workbook.save(xlsx_path)
            workbook.close()
            content = xlsx_path.read_bytes()
        rows = parse_placement_xlsx(content)
        self.assertEqual(rows[0]["sku"], 10)
        self.assertEqual(rows[0]["offer_id"], "ART")
        self.assertEqual(rows[0]["placement_cost"], 12.5)

    def test_parse_placement_xlsx_with_current_ozon_headers(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "Дата",
                "SKU",
                "Артикул",
                "Категория товара",
                "Описательный тип",
                "Склад",
                "Признак товара",
                "Суммарный объем в миллилитрах",
                "Кол-во экземпляров",
                "Платный объем в миллилитрах",
                "Кол-во платных экземпляров",
                "Начисленная стоимость размещения",
            ]
        )
        sheet.append(["04.08.2026", 10, "ART", "CAT", "TYPE", "WH", "", 12000, 3, 5000, 2, "63,5"])
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            xlsx_path = Path(tmp) / "placement.xlsx"
            workbook.save(xlsx_path)
            workbook.close()
            content = xlsx_path.read_bytes()

        rows = parse_placement_xlsx(content)

        self.assertEqual(rows[0]["sku"], 10)
        self.assertEqual(rows[0]["offer_id"], "ART")
        self.assertEqual(rows[0]["placement_cost"], 63.5)
        self.assertEqual(rows[0]["payload"]["Кол-во платных экземпляров"], 2)

    def test_placement_v25_uses_latest_nonempty_product_report_date(self) -> None:
        sql = Path("migrations/V25__ozon_placement_use_latest_nonempty_product_report.sql").read_text(encoding="utf-8")

        self.assertIn("EXISTS (\n          SELECT 1", sql)
        self.assertIn("raw.ozon_placement_report_rows", sql)
        self.assertIn("p.date_from = s.date_from", sql)
        self.assertIn("p.date_to = s.date_to", sql)

    def test_placement_v26_reads_supplies_report_by_column_position(self) -> None:
        sql = Path("migrations/V26__ozon_placement_supplies_first_paid_by_column_position.sql").read_text(encoding="utf-8")

        self.assertIn("c.column_number = 1", sql)
        self.assertIn("c.column_number = 5", sql)
        self.assertIn("c.column_number = 6", sql)
        self.assertIn("c.column_number = 8", sql)
        self.assertIn("c.column_number = 9", sql)

    def test_placement_v27_uses_ozon_stock_layer_for_first_paid_days(self) -> None:
        sql = Path("migrations/V27__ozon_placement_first_paid_uses_stock_layer.sql").read_text(encoding="utf-8")

        self.assertIn("current_stock AS", sql)
        self.assertIn("staging.ozon_stock_by_cluster", sql)
        self.assertIn("SUM(available_stock_count)", sql)
        self.assertNotIn("stock_qty) FILTER (WHERE c.column_number = 5)", sql)

    @patch("app.ops.ozon_placement_reports.OzonSellerClient")
    @patch("app.ops.ozon_placement_reports.download_report_file")
    @patch("app.ops.ozon_placement_reports.fetch_report_info")
    @patch("app.ops.ozon_placement_reports.create_placement_by_supplies_report")
    def test_download_placement_by_supplies_report_saves_xlsx_and_headers(
        self,
        create_report: Mock,
        fetch_info: Mock,
        download_file: Mock,
        _client_cls: Mock,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Поставка", "Склад", "Дней до первой платности"])
        sheet.append(["SUP-1", "WH", 3])
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            xlsx_path = Path(tmp) / "supplies.xlsx"
            workbook.save(xlsx_path)
            workbook.close()
            content = xlsx_path.read_bytes()

            create_report.return_value = ("code/1", Mock())
            fetch_info.return_value = ({"status": "success", "file": "https://example.test/supplies.xlsx"}, Mock())
            download_file.return_value = content

            result = download_placement_by_supplies_report(
                date_from=date(2026, 8, 5),
                date_to=date(2026, 8, 5),
                output_dir=Path(tmp),
                poll_attempts=1,
                poll_sleep_seconds=5,
            )

        self.assertEqual(result["code"], "code/1")
        self.assertEqual(result["rows"], 1)
        self.assertIn("Дней до первой платности", result["headers"])
        self.assertTrue(Path(result["path"]).name.startswith("ozon_placement_by_supplies_2026-08-05_2026-08-05_"))

    def test_local_file_readers(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            root = Path(tmp)
            text_path = root / "Остатки МП.txt"
            text_path.write_text("Артикул\tСМП\nART\t5\n", encoding="cp1251")
            self.assertEqual(resolve_latest_file(root, patterns=("*.txt",)), text_path)
            self.assertEqual(read_tabular_text(text_path)[0]["Артикул"], "ART")

            xlsx_path = root / "Список заказов.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Артикул", "готов"])
            sheet.append(["ART", 3])
            workbook.save(xlsx_path)
            workbook.close()
            self.assertEqual(read_first_sheet_rows(xlsx_path)[0]["готов"], 3)
            self.assertEqual(read_local_table_rows(xlsx_path)[0]["Артикул"], "ART")

    def test_powerquery_like_file_readers(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            root = Path(tmp)

            orders_path = root / "Список заказов.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "КНР"
            sheet.append(["Артикул", "согласование заказа", "в производстве", "готов", "в пути", "Дата Минск"])
            sheet.append([200, 1, 2, 3, 4, "01.07.2026"])
            sheet.append([100, 5, 6, 7, 8, "02.07.2026"])
            workbook.save(orders_path)
            workbook.close()

            pipeline = read_supply_pipeline_rows(orders_path)
            self.assertEqual([row["Артикул"] for row in pipeline], [100, 200])
            self.assertEqual(pipeline[0]["СОГЛ Заказа"], 5)
            self.assertEqual(pipeline[0]["МИНСК"], "02.07.2026")

            stocks_path = root / "Остатки МП.txt"
            stocks_path.write_text(
                "\n".join(
                    [
                        "служебная",
                        "служебная",
                        "служебная",
                        "Артикул\tДЛЯ МАРКЕТПЛЕЙСОВ\tосновной\tОтветственное хранение Великий камень\tСклад СВХ Великий камень\tТаможенный склад (Великий камень)",
                        "ед\tед\tед\tед\tед\tед",
                        "ART\t1.5\t2\t0\t0\t0",
                        "ART\t2,5\t0\t3\t0\t4",
                        "Итого\t4\t2\t3\t0\t4",
                    ]
                ),
                encoding="cp1251",
            )

            inventory = read_production_inventory_rows(stocks_path)
            self.assertEqual(inventory, [{"Артикул": "ART", "СМП": 4.0, "ОСН": 2.0, "СОХ": 3.0, "СВХ": 0.0, "ТС": 4.0}])


if __name__ == "__main__":
    unittest.main()
