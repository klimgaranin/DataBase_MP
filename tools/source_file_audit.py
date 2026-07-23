from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "local" / "audits" / "source_files"
SECRET_MARKERS = ("token", "password", "secret", "api_key", "apikey", "authorization", "private_key")


def _redact(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    if any(marker in value.lower() for marker in SECRET_MARKERS):
        return "***REDACTED***"
    return value


def _read_csv(path: Path, *, sample_rows: int) -> dict[str, Any]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                sample = fh.read(4096)
                fh.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                reader = csv.reader(fh, dialect)
                rows = []
                total_rows = 0
                for row in reader:
                    total_rows += 1
                    if len(rows) < sample_rows + 1:
                        rows.append(row)
                return {
                    "type": "csv",
                    "encoding": encoding,
                    "delimiter": dialect.delimiter,
                    "rows": total_rows,
                    "columns": len(rows[0]) if rows else 0,
                    "headers": [_redact(cell) for cell in (rows[0] if rows else [])],
                    "sample_rows": [[_redact(cell) for cell in row] for row in rows[1:]],
                }
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"CSV не прочитан: {last_error}")


def _read_xlsx(path: Path, *, sample_rows: int) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Для XLSX нужен openpyxl: python -m pip install -r requirements.txt") from exc

    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if len(rows) < sample_rows + 1:
                rows.append(list(row))
            else:
                break
        headers = rows[0] if rows else []
        sheets.append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "headers": [_redact(cell) for cell in headers],
                "sample_rows": [[_redact(cell) for cell in row] for row in rows[1:]],
            }
        )
    return {"type": "xlsx", "sheets": sheets}


def build_file_audit(path: Path, *, sample_rows: int) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        content = _read_csv(path, sample_rows=sample_rows)
    elif suffix in (".xlsx", ".xlsm"):
        content = _read_xlsx(path, sample_rows=sample_rows)
    else:
        raise ValueError("Поддерживаются только .csv, .xlsx, .xlsm")
    return {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "path": str(path),
        "file_name": path.name,
        "file_size": path.stat().st_size,
        "content": content,
    }


def write_audit(audit: dict[str, Any], output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    stem = Path(audit["file_name"]).stem.replace(" ", "_")
    json_path = output_dir / f"{stem}_audit_{stamp}.json"
    md_path = output_dir / f"{stem}_audit_{stamp}.md"
    json_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    lines = [
        "# Аудит локального файла",
        "",
        f"- Файл: {audit['file_name']}",
        f"- Размер: {audit['file_size']} bytes",
        f"- Дата: {audit['created_at']}",
        "",
    ]
    content = audit["content"]
    if content["type"] == "csv":
        lines.extend(
            [
                "## CSV",
                f"- Кодировка: {content['encoding']}",
                f"- Разделитель: `{content['delimiter']}`",
                f"- Строк: {content['rows']}",
                f"- Колонок: {content['columns']}",
                "- Заголовки: " + " | ".join(str(cell) for cell in content["headers"][:80]),
            ]
        )
    else:
        lines.append("## XLSX")
        for sheet in content["sheets"]:
            lines.append(f"### {sheet['title']}")
            lines.append(f"- Размер: {sheet['max_row']} строк, {sheet['max_column']} колонок")
            lines.append("- Заголовки: " + " | ".join(str(cell) for cell in sheet["headers"][:80]))
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return json_path, md_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit local CSV/XLSX source export")
    parser.add_argument("path", help="Path to .csv/.xlsx/.xlsm file")
    parser.add_argument("--sample-rows", type=int, default=10)
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR))
    args = parser.parse_args()

    path = Path(args.path)
    if not path.is_absolute():
        path = ROOT / path
    if not path.exists():
        raise SystemExit(f"Файл не найден: {path}")
    audit = build_file_audit(path, sample_rows=args.sample_rows)
    json_path, md_path = write_audit(audit, Path(args.output_dir))
    print(f"Файл: {path}")
    print(f"Тип: {audit['content']['type']}")
    print(f"JSON: {json_path}")
    print(f"MD: {md_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
