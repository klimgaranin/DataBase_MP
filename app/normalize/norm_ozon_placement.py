from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook


SKU_HEADERS = {"sku", "озон sku", "ozon sku", "идентификатор товара", "идентификатор товара ozon"}
OFFER_HEADERS = {"артикул", "offer_id", "offer id", "артикул продавца"}
NAME_HEADERS = {"название", "товар", "наименование", "product_name", "name"}
COST_MARKERS = ("стоимость", "размещ", "хранен", "placement", "storage", "amount", "итого")


def parse_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).replace(" ", "").replace(",", ".")))
    except (TypeError, ValueError):
        return None


def parse_decimal(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("\xa0", " ")


def parse_placement_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []

    header_idx = 0
    for idx, row in enumerate(rows[:20]):
        normalized = {_norm_header(cell) for cell in row}
        if normalized & (SKU_HEADERS | OFFER_HEADERS):
            header_idx = idx
            break

    headers = [str(cell).strip() if cell is not None else f"Column{idx}" for idx, cell in enumerate(rows[header_idx], 1)]
    normalized_headers = [_norm_header(header) for header in headers]
    result: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[header_idx + 1 :], header_idx + 2):
        payload = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        if not any(value not in (None, "") for value in payload.values()):
            continue
        result.append(
            {
                "row_number": row_number,
                "sku": _find_int(payload, normalized_headers, SKU_HEADERS),
                "offer_id": _find_text(payload, normalized_headers, OFFER_HEADERS),
                "product_name": _find_text(payload, normalized_headers, NAME_HEADERS),
                "placement_cost": _find_cost(payload, normalized_headers),
                "payload": payload,
            }
        )
    return result


def _find_text(payload: dict[str, Any], normalized_headers: list[str], names: set[str]) -> str | None:
    keys = list(payload)
    for idx, header in enumerate(normalized_headers):
        if header in names:
            value = payload.get(keys[idx])
            return str(value).strip() if value not in (None, "") else None
    return None


def _find_int(payload: dict[str, Any], normalized_headers: list[str], names: set[str]) -> int | None:
    keys = list(payload)
    for idx, header in enumerate(normalized_headers):
        if header in names:
            return parse_int(payload.get(keys[idx]))
    return None


def _find_cost(payload: dict[str, Any], normalized_headers: list[str]) -> float | None:
    keys = list(payload)
    for idx, header in enumerate(normalized_headers):
        if any(marker in header for marker in COST_MARKERS):
            parsed = parse_decimal(payload.get(keys[idx]))
            if parsed is not None:
                return parsed
    return None
