from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


@dataclass(frozen=True)
class ExcelTableData:
    name: str
    sheet_name: str
    ref: str
    rows: list[dict[str, Any]]


def read_excel_tables(path: str | Path, *, table_names: set[str] | None = None) -> list[ExcelTableData]:
    """
    Read materialized Excel tables from .xlsx/.xlsm without refreshing Power Query
    and without executing VBA macros.
    """
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=False, keep_vba=False)
    try:
        result: list[ExcelTableData] = []
        for sheet in workbook.worksheets:
            for table_name in sheet.tables:
                table = sheet.tables[table_name]
                if table_names is not None and table_name not in table_names:
                    continue
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                values = list(
                    sheet.iter_rows(
                        min_row=min_row,
                        max_row=max_row,
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True,
                    )
                )
                if not values:
                    rows: list[dict[str, Any]] = []
                else:
                    headers = [str(cell).strip() if cell is not None else f"Column{idx}" for idx, cell in enumerate(values[0], 1)]
                    rows = []
                    for raw_row in values[1:]:
                        item = {headers[idx]: raw_row[idx] for idx in range(len(headers))}
                        if any(value not in (None, "") for value in item.values()):
                            rows.append(item)
                result.append(ExcelTableData(name=table_name, sheet_name=sheet.title, ref=table.ref, rows=rows))
        return result
    finally:
        workbook.close()
