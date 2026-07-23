from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def resolve_latest_file(path: str | Path, *, patterns: tuple[str, ...]) -> Path:
    source = Path(path)
    if source.is_file():
        return source
    if not source.exists():
        raise FileNotFoundError(f"Источник не найден: {source}")
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(item for item in source.glob(pattern) if item.is_file())
    if not candidates:
        raise FileNotFoundError(f"В папке {source} не найдено файлов: {patterns}")
    return max(candidates, key=lambda item: item.stat().st_mtime)


def resolve_latest_file_preferred(path: str | Path, *, patterns: tuple[str, ...]) -> Path:
    source = Path(path)
    if source.is_file():
        return source
    if not source.exists():
        raise FileNotFoundError(f"Источник не найден: {source}")
    for pattern in patterns:
        candidates = [item for item in source.glob(pattern) if item.is_file()]
        if candidates:
            return max(candidates, key=lambda item: item.stat().st_mtime)
    raise FileNotFoundError(f"В папке {source} не найдено файлов: {patterns}")


def read_excel_sheet_rows(path: str | Path, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    return rows_to_dicts(rows)


def read_first_sheet_rows(path: str | Path) -> list[dict[str, Any]]:
    return read_excel_sheet_rows(path)


def rows_to_dicts(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    header_idx = _find_header_row(rows)
    headers = [str(cell).strip() if cell is not None else f"Column{idx}" for idx, cell in enumerate(rows[header_idx], 1)]
    result = []
    for raw_row in rows[header_idx + 1 :]:
        row = {headers[idx]: raw_row[idx] if idx < len(raw_row) else None for idx in range(len(headers))}
        if any(value not in (None, "") for value in row.values()):
            result.append(row)
    return result


def read_supply_pipeline_rows(path: str | Path) -> list[dict[str, Any]]:
    rows = read_excel_sheet_rows(path, sheet_name="КНР")
    selected = []
    for row in rows:
        item_article = row.get("Артикул")
        if item_article in (None, ""):
            continue
        selected.append(
            {
                "Артикул": item_article,
                "СОГЛ Заказа": row.get("согласование заказа"),
                "В ПРОИЗВ": row.get("в производстве"),
                "ГОТОВ": row.get("готов"),
                "В ПУТИ": row.get("в пути"),
                "МИНСК": row.get("Дата Минск"),
            }
        )
    return sorted(selected, key=lambda row: str(row.get("Артикул") or ""))


def read_production_inventory_rows(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    if source.suffix.lower() == ".txt":
        return _read_production_inventory_txt(source)
    rows = read_local_table_rows(source)
    return _transform_production_inventory_rows(rows, remove_last=False)


def read_local_table_rows(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_first_sheet_rows(source)
    if suffix == ".xls":
        return read_xls_rows(source)
    if suffix in {".txt", ".csv", ".tsv"}:
        delimiter = ";" if suffix == ".csv" else "\t"
        return read_tabular_text(source, delimiter=delimiter)
    raise ValueError(f"Неподдерживаемый формат файла: {source}")


def _read_production_inventory_txt(path: Path) -> list[dict[str, Any]]:
    rows = _read_tabular_text_rows(path, delimiter="\t")
    if len(rows) <= 4:
        return []
    promoted_rows = rows[3:]
    headers = [str(cell).strip() if cell is not None else f"Column{idx}" for idx, cell in enumerate(promoted_rows[0], 1)]
    data_rows = []
    for raw_row in promoted_rows[2:]:
        row = {headers[idx]: raw_row[idx] if idx < len(raw_row) else None for idx in range(len(headers))}
        if any(value not in (None, "") for value in row.values()):
            data_rows.append(row)
    return _transform_production_inventory_rows(data_rows, remove_last=True)


def _transform_production_inventory_rows(rows: list[dict[str, Any]], *, remove_last: bool) -> list[dict[str, Any]]:
    rename = {
        "ДЛЯ МАРКЕТПЛЕЙСОВ": "СМП",
        "основной": "ОСН",
        "Ответственное хранение Великий камень": "СОХ",
        "Склад СВХ Великий камень": "СВХ",
        "Таможенный склад (Великий камень)": "ТС",
    }
    rows_to_use = rows[:-1] if remove_last and rows else rows
    grouped: dict[str, dict[str, Any]] = {}
    for source_row in rows_to_use:
        article = _clean_article(source_row.get("Артикул"))
        if not article:
            continue
        normalized = {"Артикул": article}
        for source_col, target_col in rename.items():
            normalized[target_col] = _parse_ru_number(source_row.get(source_col, source_row.get(target_col)))
        if not any(normalized[col] for col in ("СМП", "ОСН", "СОХ", "СВХ", "ТС")):
            continue
        target = grouped.setdefault(article, {"Артикул": article, "СМП": 0.0, "ОСН": 0.0, "СОХ": 0.0, "СВХ": 0.0, "ТС": 0.0})
        for col in ("СМП", "ОСН", "СОХ", "СВХ", "ТС"):
            target[col] += normalized[col]
    return list(grouped.values())


def read_xls_rows(path: str | Path) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Для чтения .xls нужен xlrd: python -m pip install -r requirements.txt") from exc

    workbook = xlrd.open_workbook(str(path))
    sheet = workbook.sheet_by_index(0)
    rows = [tuple(sheet.row_values(row_idx)) for row_idx in range(sheet.nrows)]
    return rows_to_dicts(rows)


def _read_tabular_text_rows(path: str | Path, *, delimiter: str = "\t", encodings: tuple[str, ...] = ("cp1251", "utf-8-sig", "utf-8")) -> list[tuple[Any, ...]]:
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with Path(path).open("r", encoding=encoding, newline="") as fh:
                return [tuple(row) for row in csv.reader(fh, delimiter=delimiter)]
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Текстовый файл не прочитан: {path}: {last_error}")


def read_tabular_text(path: str | Path, *, delimiter: str = "\t", encodings: tuple[str, ...] = ("cp1251", "utf-8-sig", "utf-8")) -> list[dict[str, Any]]:
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with Path(path).open("r", encoding=encoding, newline="") as fh:
                reader = csv.DictReader(fh, delimiter=delimiter)
                return [dict(row) for row in reader if any(value not in (None, "") for value in row.values())]
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Текстовый файл не прочитан: {path}: {last_error}")


def _clean_article(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_ru_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(" ", "").replace(".", ",").replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    for idx, row in enumerate(rows[:30]):
        normalized = {str(cell or "").strip().lower() for cell in row}
        if "артикул" in normalized or "ваш sku" in normalized:
            return idx
    return 0
